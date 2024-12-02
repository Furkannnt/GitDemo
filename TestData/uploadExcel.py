import time
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

def update_excel_data(filePath, searchTerm, colName, newValue):
    # Belirtilen dosya yolundan Excel dosyasını yükle
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.active
    # Verilerin depolanacağı sözlük
    data_dict = {}

    # Kolon adını bul
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            data_dict["col"] = i

    # Arama terimini bul ve satır numarasını al
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                data_dict["row"] = i

    # Yeni değeri hücreye ata
    sheet.cell(row=data_dict["row"], column=data_dict["col"]).value = newValue
    # Değişiklikleri kaydet
    workbook.save(filePath)

# Dosya yolunu belirle
file_path = r"C:\Users\PROVEN-LAPTOP\Downloads\download.xlsx"
# WebDriver'ı başlat
driver = webdriver.Chrome()
driver.implicitly_wait(5)
# Web sitesine git
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()

# İndirme butonuna tıkla
driver.find_element(By.ID, "downloadButton").click()
# Dosyanın indirilmesi için zaman tanıyın
time.sleep(3)

# Excel dosyasını düzenle
update_excel_data(file_path, "Apple", "price", "999")

# Dosyayı yükle
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

# Bildirim mesajının görünmesini bekleyin
wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(EC.visibility_of_element_located(toast_locator))
info = driver.find_element(*toast_locator).text
# Başarı mesajını doğrula
assert "Updated" in info

# Tarayıcının manuel olarak kapatılmasını sağla
print("Test tamamlandı, tarayıcıyı kapatmak için manuel olarak kapatın.")
input("Devam etmek için Enter tuşuna basın...")
