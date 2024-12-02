import openpyxl
import pytest

class TestExcelOperations:

    @pytest.fixture(scope="class")
    def create_document(self):
        # Excel dosyasını oluşturun
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        for j in range(10):
            sheet.cell(row=1, column=j+1, value="")  # Hücreleri oluştur
        workbook.save("C:/Users/PROVEN-LAPTOP/Desktop/pythonAddData.xlsx")
        yield workbook

    def test_write_numbers(self, create_document):
        # Belirtilen sayıları ilk satıra yazma
        sheet = create_document.active  # İlk tabloyu seçiyoruz
        for j in range(10):
            sheet.cell(row=1, column=j+1, value=str(j + 1))  # Sütunları sırasıyla seçiyoruz ve veri ekliyoruz
        create_document.save("C:/Users/PROVEN-LAPTOP/Desktop/pythonAddData.xlsx")

    def test_read_numbers(self):
        # Dosyayı yeniden yükleme ve okuma
        workbook = openpyxl.load_workbook("C:/Users/PROVEN-LAPTOP/Desktop/pythonAddData.xlsx")
        sheet = workbook.active  # İlk tabloyu seçiyoruz
        for j in range(10):
            cell = sheet.cell(row=1, column=j+1)
            cell_value = cell.value
            assert cell_value == str(j + 1), f"Hücre değeri beklenen değerden farklı: {cell_value}"
